"""
Genera el reporte Excel.
"""
 
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
 
 
def generar_excel(df: pd.DataFrame, alertas: list[str]) -> Path:
    """Crea un .xlsx con tabla de precios + hoja de alertas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios"
 
    # ── Paleta de colores ──────────────────────────────────────────────
    AZUL_HEADER = "1F4E79"
    AZUL_CLARO  = "D6E4F0"
    VERDE       = "E2EFDA"
    ROJO        = "FFE0E0"
    GRIS_FILA   = "F5F5F5"
 
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    # ── Título ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Reporte Dólar — {datetime.now():%d/%m/%Y %H:%M}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = PatternFill("solid", fgColor=AZUL_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
 
    # ── Encabezados ────────────────────────────────────────────────────
    headers = ["Tipo", "Compra ($)", "Venta ($)", "Spread ($)", "Variación"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border
 
    # ── Filas de datos ─────────────────────────────────────────────────
    for i, row in enumerate(df.itertuples(), start=3):
        spread = round(row.venta - row.compra, 2)
        valores = [row.tipo, row.compra, row.venta, spread, row.variacion]
 
        fill_color = AZUL_CLARO if row.tipo == "Blue" else (GRIS_FILA if i % 2 == 0 else "FFFFFF")
        if row.tipo == "Blue" and row.venta > 1500:
            fill_color = ROJO
 
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
            if col in (2, 3, 4):
                cell.number_format = '"$"#,##0.00'
 
    # ── Ancho de columnas ──────────────────────────────────────────────
    anchos = [16, 14, 14, 14, 14]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    # ── Hoja de alertas (si hay) ───────────────────────────────────────
    if alertas:
        ws_a = wb.create_sheet("Alertas")
        ws_a["A1"] = "Alertas generadas"
        ws_a["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        ws_a["A1"].fill = PatternFill("solid", fgColor="C00000")
        ws_a.column_dimensions["A"].width = 50
        for fila, msg in enumerate(alertas, start=2):
            ws_a.cell(row=fila, column=1, value=msg)
 
    # ── Guardar ────────────────────────────────────────────────────────
    output_path = Path(f"reporte_dolar_{datetime.now():%Y%m%d_%H%M}.xlsx")
    wb.save(output_path)
    print(f"  Excel guardado: {output_path}")
    return output_path